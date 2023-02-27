
print("")
print("        __|__     ")
print("   ---o--(_)--o---")
print("          _   ____     ___    ___    ____    ____")
print("         | | |  _ \   / __|  / _ \  |  __|  / _  |")
print("         | | | | | | | (__  | (_) | | |    | (_| |")
print("         |_| |_| |_|  \___|  \___/  |_|    \___,_|")

print("\n       WELCOME TO INCORA'S OASIS WEB SCRAPER!")
print(" ")


import sys
import datetime
import time
import schedule
import os
import glob
import pathlib
from pathlib import Path
import shutil

import openpyxl
import pandas as pd
from functions import locate_col
import math
import warnings
warnings.simplefilter("ignore")  # used to avoid user warnings when opening excel file

import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait


def app_schedule():
    # GET USER DATA
    user_path = str(Path.home() / "Downloads")

    path = (user_path + '/*.pdf')  # * means all if need specific format then *.pdf

    # LAUNCH WEBSITE
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-notifications")

    try:
        URL = 'https://www.iaqg.org/oasis/mywatchlist'
        driver = webdriver.Chrome(options=options)
        driver.get(URL)
        page = requests.get(URL)
        soup = BeautifulSoup(page.content, 'html.parser')
    except Exception:
        print('\n   ERROR: WEB BROWSER NOT COMPATIBLE WITH DRIVER\n')
        print('\n   TRY AGAIN AFTER REPLACING THE DRIVER FILE...\n')
        time.sleep(10)
        sys.exit()

    # ENABLE COOKIES
    cookies = driver.find_element(By.ID, "onetrust-accept-btn-handler").click()

    # TRY TO LOG IN

    try:
        userID = driver.find_element(By.ID, "frm-login-1-input")
        userID.click()
        userID.send_keys(' USER LOGIN ')
        password = driver.find_element(By.ID, "frm-login-2-input")
        password.click()
        password.send_keys(' USER PASSWORD ')
        login = driver.find_element(By.CSS_SELECTOR, "#frm-login-3 > div.frmi-s > a > div > div > span").click()
        time.sleep(2)
    except:
        print("Provided login or password is incorrect!")
        time.sleep(5)
        sys.exit()

    # DOWNLOAD WATCHLIST FILE FROM OASIS
    download_tab = driver.find_element(By.CSS_SELECTOR, '#downloadwatchlist > span > em').click()
    choose_button = driver.find_element(By.CSS_SELECTOR, '#frm-dwn-1 > div.frmi-c.frmi-c-one > div > ul > li:nth-child(2) > div.frmi-c-l > label').click()
    download_watchlist_button = driver.find_element(By.CSS_SELECTOR, '#download-watchlist > div > div > span > input[type=submit]').click()

    # LOCATE DOWNLOADED FILE
    time.sleep(5)  # wait for file to download
    path = (user_path + '/*.xlsx')  # * means all if need specific format then *.pdf
    list_of_files = glob.glob(path)
    watchlist_file = max(list_of_files, key=os.path.getctime)
    watchlist_old_path = os.path.abspath(watchlist_file)

    # OPEN EXCEL FILE
    print("Loading Excel file ...")

    try:
        excel = openpyxl.load_workbook(watchlist_file)
        print("\n Success! Excel file is loaded.")
        print("Now let me look through the file... \n")
    except Exception:
        print("Cannot open Excel file, please make sure its extension is xlsx or xlsm...")
        time.sleep(10)
        sys.exit()


    xl_workbook = pd.ExcelFile(watchlist_file)  # Load the excel workbook
    df = xl_workbook.parse("My WatchList Supplier")  # Parse the sheet into a dataframe

    # Cast the desired column into a python list
    OIN_list = []
    expiry_date_list = []
    status_list = []
    result_list = []

    # CHECK FOR EXPIRED DATES & LOOP THROUGH OIN NUMBERS


    for i in range(1048575):  # 1,048,576 is max number of rows in excel
        supplier_name_col_item = excel['My WatchList Supplier'].cell(row=2 + i, column=2).value
        if supplier_name_col_item is None or supplier_name_col_item == '':
            break
        OIN_col_item = excel['My WatchList Supplier'].cell(row=2 + i, column=1).value
        expiry_date_col_item = excel['My WatchList Supplier'].cell(row=2 + i, column=14).value
        if expiry_date_col_item is not None and expiry_date_col_item != '':
            days_since1 = math.floor((datetime.datetime.today() - expiry_date_col_item).total_seconds() / 86400.00)  # seconds in a day, round down
            if days_since1 > 3:  # 3 days extra for OASIS to update supplier info
                OIN_list.append(math.floor(OIN_col_item))
                expiry_date_list.append(expiry_date_col_item)


    if not OIN_list:  #  if list is empty
        print("\n NOTHING TO SCRAP! ALL CERTS ARE UP TO DATE. EXITING...")
        time.sleep(3)
        sys.exit()
    else:
        #  CREATE DESTINATION FOLDER TO KEEP REPORTS
        cd = datetime.datetime.now().strftime('%Y-%m-%d %H;%M;%S')  # get current datetime
        parent_dir = (r' PATH TO FOLDER ON COMPANY SHARED DRIVE')
        dest_dir = os.path.join(parent_dir, cd)
        dd = os.mkdir(dest_dir)

        # Calculate workload & print summary
        processing_time = round(((len(list(filter(None, OIN_list)))) * 22) / 60)  # minutes

        print("\n\n   SUMMARY:")
        print("\n     NUMBER OF SEARCHES TO COMPLETE:", len(list(filter(None, OIN_list))))
        print("\n     POTENTIAL PROCESSING TIME:", str(processing_time), "MINUTES")
        print("\n\n   STARTING WEB SEARCH... \n\n")


    # locate elements
    supplier_col = driver.find_elements(By.XPATH, "//*[@id='csd--suppliers-list']/div[3]/table/tbody/tr[1]/td[1]/div[1]/div[3]/span/a/span/strong")
    status_col = driver.find_elements(By.XPATH, "//*[@id='csd--suppliers-list']/div[3]/table/tbody/tr[1]/td[2]/div[1]")

    # cast lists to hold results
    supplier_results = []
    status_results = []

    # GET Certified Suppliers Directory PAGE
    dataSearch_tab = driver.find_element(By.CSS_SELECTOR, '#hdr-main-data > a').click()
    certifiedSuppDir_tab = driver.find_element(By.CSS_SELECTOR, '#hdr-main-data > div > div.omega > ul > li:nth-child(1) > a').click()
    time.sleep(2)

    # GET CURRENT TIME TO COUNT TIME
    start_time = time.time()

    # START SCRAPING

    for OIN in OIN_list:
        try:
            OIN_field = driver.find_element(By.ID, "supplierOIN")
            OIN_field.click()
            OIN_field.clear()
            OIN_field.send_keys(OIN)
            OIN_field.send_keys(Keys.ENTER)
            time.sleep(2)
            try:
                if driver.find_element(By.ID, "popmsg-c").is_displayed():
                    status_results.append("No matching suppliers were found")
                    supplier_results.append("Not found")
                    continue
                else:
                    supplier_name = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                        By.XPATH, '//*[@id="csd--suppliers-list"]/div[3]/table/tbody/tr/td[1]/div[1]/div[3]/span/a/span/strong'
                                    ))).text
                    OIN_id = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                        By.XPATH, '//*[@id="csd--suppliers-list"]/div[3]/table/tbody/tr/td[1]/div[1]/div[1]/em[1]/span'
                                    ))).text
                    status_of_cert = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                        By.XPATH, '//*[@id="csd--suppliers-list"]/div[3]/table/tbody/tr/td[2]/div[1]'
                                    ))).text
                    supplier_results.append(supplier_name)
                    status_results.append(status_of_cert)
                    supplier_name_button = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((
                                                By.CSS_SELECTOR, '#csd--suppliers-list > div.tbl-container > table > tbody > tr > td:nth-child(1) > div.ct-h.ct-h-basic > div.ct-org > span > a'
                                            )))
                    supplier_name_button.click()
                    time.sleep(2)
                    try:
                        if driver.find_element(By.ID, "overlay-wrapper").is_displayed():
                            driver.find_element(By.ID, "yes-btn").click()
                    except:
                        pass
                    expire_field = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                        By.XPATH, '//*[@id="frm-aud-21"]/div[2]/div/ul/li/span'
                                    ))).text
                    replaced_expire_date = expire_field.replace("/", "")
                    download_button = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((
                                            By.CSS_SELECTOR, '#tbl-uploaded > div.tbl-container > table > tbody > tr:nth-child(3) > td > a'
                                        )))
                    download_button.click()
                    time.sleep(5)
                    # RENAME DOWNLOADED FILE
                    name_element = driver.find_element(By.CSS_SELECTOR, "#watchlist-toggle > strong").text
                    path = (user_path + '/*.pdf')  # * means all if need specific format then *.pdf
                    list_of_files = glob.glob(path)
                    latest_file = max(list_of_files, key=os.path.getctime)
                    old_file_path = os.path.abspath(latest_file)
                    # REMOVE FORBIDDEN CHARACTERS
                    char_remov = ["/", "|", ":", "*", "?", "!", "<", ">"]
                    for char in char_remov:
                        if char in name_element:
                            new_name = name_element.replace(char, " ")
                            # print(new_name)  <-- sanity check
                            break
                        else:
                            new_name = name_element
                    new_name_path = (user_path + '\\' + OIN_id + '_' + new_name + '_' + replaced_expire_date + '.pdf')
                    print("new_name_path is ", new_name_path)
                    time.sleep(3)
                    os.rename(old_file_path, new_name_path)
                    time.sleep(3)
                    # MOVE DOWNLOADED FILE
                    shutil.move(new_name_path, dest_dir)
                    time.sleep(3)
                    driver.get('https://www.iaqg.org/oasis/csd')
                    time.sleep(3)
            except FileExistsError:
                os.remove(latest_file)
                driver.get('https://www.iaqg.org/oasis/csd')
                time.sleep(3)
                continue
            except (shutil.Error, FileNotFoundError) as error:
                driver.get('https://www.iaqg.org/oasis/csd')
                time.sleep(3)
                continue
        except TypeError:
            continue

    # end web scraping
    driver.quit()  # close website
    print("\n WEB SEARCH COMPLETED!")

    #  CALCULATE TIME SPENT ON WEB SCRAPING
    end_time = time.time()
    print("\n TOOK %s MINUTES TO SCRAP RESULTS FROM OASIS" % (round((end_time - start_time) / 60)))

    # save data to excel
    print("\n PREPARING REPORT...")

    df_OIN = pd.DataFrame(
        {
            'Supplier name': supplier_results,
            'OIN #' : OIN_list,
            'Status': status_results,
            'Expiry date': expiry_date_list,
        }
    )

    # CREATE XLSX REPORT --> one file, divided by sheets

    date_string = datetime.datetime.now().strftime('%Y%m%d %H;%M;%S')
    with pd.ExcelWriter("Scrapped Results_" + date_string +".xlsx") as writer:
        df_OIN.to_excel(writer, sheet_name="Sheet1", index=False)

    # MOVE REPORT

    filename = "Scrapped Results_" + date_string +".xlsx"
    filename_path = os.path.abspath(filename)
    shutil.move(filename_path, dest_dir)

    # MOVE WATCHLIST
    try:
        shutil.move(watchlist_old_path, dest_dir)
    except PermissionError:
        pass


    print("\n All files downloaded successfuly!")

    time.sleep(5)  # freeze time for 10 secs to enable reading above info

# Every friday at 12:00 app_schedule() is called
schedule.every(4).weeks.at("08:00").do(app_schedule)


# Loop so that the scheduling task
# keeps on running all time.

while True:

    # Checks whether a scheduled task
    # is pending to run or not
    schedule.run_pending()
    time.sleep(1)