# -*- coding: utf-8 -*-
"""
Created on Thu Jan 18 20:00:57 2024

@author: klecznar
"""

import os, sys, math, time, glob, shutil, logging
from pathlib import Path

import openpyxl
import pandas as pd
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import TimeoutException, SessionNotCreatedException, NoSuchElementException

# OIN - OASIS Identification Number

# WEBSITE SETUP
options = webdriver.ChromeOptions()
# options.add_argument('--headless=new')
options.add_argument("--disable-popup-blocking")
options.add_argument("--disable-notifications")

# PATH TO OASIS WEBSITE
URL = 'https://oasis.iaqg.org/search/intact-frontend/suppliers'

# list to hold data from Excel
OIN_list = []
OIN_found_list = []
OIN_not_found_list = []

# lists to hold web elements
supplier_name_results = []
supplier_address_results = []
contact_results = []
email_results = []
status_results = []
series_results = []
standard_results = []
latest_certificate_results = []
latest_audit_results = []
expiry_results = []
scope_results = []

standard_results2 = []
expiry_results2 = []
status_results2 = []
scope_results2 = []
email_results2 = []


def get_credentials():
    
    """
    Function asks user to provide credentials to the OASIS website
    """
    
    global login, passcode
    
    login = input("Please provide login to OASIS: ")
    passcode = input("Please provide password to OASIS: ")
    
    return login, passcode


def get_expired_OINs():
    
    """
    Function used to get only OIN numbers with expired certification status
    """

    try:
        excel = openpyxl.load_workbook('QMS Certifications.xlsm')
        print("Let me look through the file...")
        
        # CHECK FOR EXPIRED DATES & GET OIN NUMBERS
    
        for i in range(1048575):  # 1,048,576 is max number of rows in excel
            # get supplier name starting from row 3
            supplier_name_col_item = excel['Database'].cell(row=3 + i, column=1).value
            # break out of loop if cell is empty
            if supplier_name_col_item is None or supplier_name_col_item == '':
                break
            # get OIN starting from row 3
            OIN_col_item_1 = excel['Database'].cell(row=3 + i, column=6).value
            OIN_col_item_2 = excel['Database'].cell(row=3 + i, column=7).value
            # get expiry date starting from row 3
            expiry_date_col_item = excel['Database'].cell(row=3 + i, column=15).value
            if expiry_date_col_item is not None and expiry_date_col_item != '':
                # if TypeError occurs, convert the str type date to datetime type date
                # 86400 seconds in a day, rounds down the difference
                try:
                   conv_expiry_date = datetime.strptime(expiry_date_col_item, '%d/%m/%Y')
                   days_since = math.floor((datetime.today() - conv_expiry_date).total_seconds() / 86400.00)
                except:
                    days_since = math.floor((datetime.today() - expiry_date_col_item).total_seconds() / 86400.00)
                if days_since > 3:   # 3 days extra for OASIS to update supplier info
                    if OIN_col_item_1 is not None:  
                        OIN_list.append(OIN_col_item_1)
                    if  OIN_col_item_2 is not None:
                        OIN_list.append(OIN_col_item_2)       
    except Exception as a:
        logging.exception(a)
        time.sleep(10)
        sys.exit()
        

    return OIN_list


def get_selected_OINs():
    
    """
    Function used to get all OIN numbers which user is interested in
    from a specific 'ALL.xlsx' file
    """
    
    try:
        excel = openpyxl.load_workbook("ALL.xlsx")
        
        for i in range(1048575):  # 1,048,576 is max number of rows in excel
            OIN_col_item = excel['Sheet1'].cell(row=1 + i, column=1).value
            if OIN_col_item is None or OIN_col_item == '':
                break
            if OIN_col_item is not None and OIN_col_item != '' and len(str(OIN_col_item)) == 10:
                OIN_list.append(OIN_col_item)
                        
    except Exception:
        logging.exception("Something went wrong!")
        time.sleep(10)
        sys.exit()
        
    
    return OIN_list



def alert():
    
    """
    Function prompts alert to user with following data: 
        number of OIN numbers to be scraped from the OASIS website
        and how long the process will take
    It also creates a destination folder to keep all downloaded certificates
    and a report
    """
    
    # CALCULATE WORKLOAD AND PROMPT ALERT TO USER
    
    if not OIN_list:  #  if list is empty
        print("NOTHING TO SCRAPE! ALL CERTS ARE UP TO DATE. EXITING...")
        time.sleep(3)
        sys.exit()
    else:
        #  CREATE DESTINATION FOLDER TO STORE CERTIFICATES 
        # get current datetime
        cd = datetime.now().strftime('%Y-%m-%d %H;%M;%S')
        parent_dir = r'\\components\DATA\PLS-Public\PLSQA\OASIS\OASIS_QMS certificates'
        destination_dir_path = os.path.join(parent_dir, cd)
        # create folder 
        os.mkdir(destination_dir_path)
        
        print("")
        print("SUMMARY: ")
        print("NUMBER OF SEARCHES TO COMPLETE: ", len(OIN_list))
        print("POTENTIAL PROCESSING TIME:", str(len(OIN_list)), "MINUTES")
        print("STARTING WEB SEARCH...")
        print("")
        
    return destination_dir_path


def log_in(driver):
    
    """
    Function uses provided credentials in the get_credentials() function
    and gets access to the OASIS website
    """

    
    # LOG IN
    userID = driver.find_element(By.ID, "txtUser")
    userID.click()
    userID.send_keys('watchlist@pattonair.com')
    password = driver.find_element(By.ID, "txtPassword")
    password.send_keys('Pattonair0321')
    password.send_keys(Keys.ENTER)
    time.sleep(1)
    

def download_certs():
 
    """
    Function used to web scrap data, download certificates
    """
    
    # PROMPT FUNCTION AND GET VARIABLE
    destination_dir_path = alert()
    

    # GET CURRENT TIME TO COUNT TIME
    start_time = time.time()

    # START SCRAPING

    for OIN in OIN_list:
        try:
            driver = webdriver.Chrome(options=options)            
            driver.minimize_window()
            driver.get(URL)
            time.sleep(1)
            # LOG IN
            log_in(driver)
            
            # GET TO DESIRED PAGE    
            searchTrack_button = driver.find_element(By.CSS_SELECTOR, '#WucKoLandingPage > div.col-group.col6-12.module-selection > div.module-selection__card-wrapper > div:nth-child(1) > div')
            searchTrack_button.click()
            time.sleep(1)
            moreOptions_button = WebDriverWait(driver, 20).until(EC.presence_of_element_located((
                    By.CSS_SELECTOR, 'body > intact-root > ngx-intact-sidebar-menu > mat-sidenav-container > mat-sidenav-content > div > intact-suppliers > ngx-intact-advanced-search > mat-card > form > div:nth-child(2) > button > span.mat-button-wrapper > mat-icon')))
            moreOptions_button.click()
            OIN_field = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((
                By.ID, "mat-input-1")))
            OIN_field.click()
            OIN_field.clear()
            OIN_field.send_keys(OIN)
            OIN_field.send_keys(Keys.ENTER)
        
            # GET TO DESIRED SUPPLIER
            row = WebDriverWait(driver, 20).until(EC.presence_of_element_located((
                    By.CSS_SELECTOR, 'body > intact-root > ngx-intact-sidebar-menu > mat-sidenav-container > mat-sidenav-content > div > intact-suppliers > ngx-intact-table > table > tbody > tr')))
            row.click()
            time.sleep(2)
            
            # if a pop-up with Level 2 Data Access warning gets displayed then handle it and continue
            try:
                level2 = WebDriverWait(driver, 20).until(EC.presence_of_element_located((
                        By.XPATH, '//*[@id="cdk-overlay-0"]')))
                cont = WebDriverWait(driver, 20).until(EC.presence_of_element_located((
                        By.CSS_SELECTOR, "#mat-dialog-0 > intact-level2-data-terms > ngx-intact-overlay > div > div.ngx-intact-overlay_footer > div > div > button.mat-focus-indicator.mr-3.mat-button.mat-button-base.mat-primary")))
                if level2.is_displayed():
                    cont.click()
                    time.sleep(1)
            except:
                pass
            
            # condition to add supplier to the watchlist if not already added
            try:
                add_to_watchlist = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                    By.XPATH, '/html/body/intact-root/div/intact-supplier-detail/div/mat-card/div/div[1]/div/intact-company-info/div/div[2]/intact-watchlist-assign/div/div/button'
                                )))
                if add_to_watchlist.text == 'ADD TO WATCHLIST':
                    add_to_watchlist.click()
            except:
                pass
            
            # get supplier data from website 
            # download certificate and create report
            supplier_name = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                By.XPATH, '/html/body/intact-root/div/intact-supplier-detail/div/mat-card/div/div[1]/div/intact-company-info/div/h2'
                            ))).text
            supplier_address = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                By.XPATH, '/html/body/intact-root/div/intact-supplier-detail/div/mat-card/div/div[1]/div/intact-company-info/div/div[1]/div/div[1]/div[2]'
                            ))).text
            contact_person = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                By.XPATH, '/html/body/intact-root/div/intact-supplier-detail/div/mat-card/div/div[1]/div/intact-company-info/div/div[1]/div/div[2]/div[1]/p'
                            ))).text
            email = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                By.XPATH, '/html/body/intact-root/div/intact-supplier-detail/div/mat-card/div/div[1]/div/intact-company-info/div/div[1]/div/div[2]/div[4]/a'
                            ))).text
            cert_status = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                By.XPATH, '//*[@id="mat-tab-content-0-0"]/div/div/intact-certificate-details/div/div[3]/div[4]/div/intact-status-chip'
                            ))).text
            series = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                By.XPATH, '/html/body/intact-root/div/intact-supplier-detail/div/div/div[1]/mat-card[1]/div/div[2]/div[2]'
                            ))).text
            standard = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                By.XPATH, '/html/body/intact-root/div/intact-supplier-detail/div/div/div[1]/mat-card[1]/div/div[2]/div[3]'
                            ))).text
            latest_cert = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                By.XPATH, '/html/body/intact-root/div/intact-supplier-detail/div/div/div[1]/mat-card[1]/div/div[2]/div[4]'
                            ))).text
            latest_audit = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                By.XPATH, '/html/body/intact-root/div/intact-supplier-detail/div/div/div[1]/mat-card[1]/div/div[2]/div[5]'
                            ))).text
            expiry_date = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                By.XPATH, '//*[@id="mat-tab-content-0-0"]/div/div/intact-certificate-details/div/div[3]/div[2]'
                            ))).text
            stripped_expiry_date = expiry_date.strip('Expiry date')
            scope = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                                By.XPATH, '//*[@id="mat-tab-content-0-0"]/div/div/intact-certificate-details/div/div[4]/div[6]/p'
                            ))).text
            
            OIN_found_list.append(OIN)
            supplier_name_results.append(supplier_name)
            supplier_address_results.append(supplier_address)
            contact_results.append(contact_person)
            email_results.append(email)
            status_results.append(cert_status)
            series_results.append(series) 
            standard_results.append(standard) 
            latest_certificate_results.append(latest_cert)
            latest_audit_results.append(latest_audit)
            expiry_results.append(stripped_expiry_date)
            scope_results.append(scope)
            
            standard_results2.append(standard)
            expiry_results2.append(expiry_date)
            status_results2.append(cert_status)
            scope_results2.append(scope)
            email_results2.append(email)

            # download certificate
            download_button = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((
                                    By.CSS_SELECTOR, '#mat-tab-content-0-0 > div > div > intact-certificate-details:nth-child(1) > div > div.item-header > button > span.mat-button-wrapper'
                                )))
            download_button.click()
            time.sleep(3)
            # rename and move downloaded file
            file_handling(driver, OIN, stripped_expiry_date, destination_dir_path)
            time.sleep(1)
            driver.quit()
        # IF OIN/SUPPLIER NOT FOUND - error raised due to web element not being located
        except TimeoutException:
            OIN_not_found_list.append(OIN)
            supplier_name_results.append("No matches for given search criteria")
            supplier_address_results.append(" ")
            contact_results.append(" ")
            email_results.append(" ")
            status_results.append(" ")
            series_results.append(" ") 
            standard_results.append(" ") 
            latest_certificate_results.append(" ")
            latest_audit_results.append(" ")
            expiry_results.append(" ")
            scope_results.append(" ")
            
            driver.quit()
            continue
        except NoSuchElementException:
            logging.exception("Provided login or password is incorrect!")
            time.sleep(5)
            sys.exit()
        except SessionNotCreatedException:
            logging.exception('ERROR: WEB BROWSER NOT COMPATIBLE WITH DRIVER')
            logging.exception('TRY AGAIN AFTER REPLACING THE DRIVER FILE...n')
            time.sleep(10)
            sys.exit()


    # END WEB SCRAPING
    print("WEB SEARCH COMPLETED!")

    #  CALCULATE TIME SPENT ON WEB SCRAPING
    end_time = time.time()
    print("TOOK %s MINUTES TO SCRAPE RESULTS FROM OASIS" % (round((end_time - start_time) / 60)))    
    
    #CREATE REPORT
    create_report(destination_dir_path)
    
    return supplier_name_results, supplier_address_results, contact_results, email_results
    return status_results, series_results, standard_results, latest_certificate_results,
    return latest_audit_results, expiry_results, scope_results


def file_handling(driver, OIN, stripped_expiry_date, destination_dir_path):
    
    """
    Function used to rename and move downloaded certificates to desired folder
    """
    
    # GET PATH TO USER'S DOWNLOADS DIRECTORY
    user_path = str(Path.home() / "Downloads")
    path = (user_path + '/*.pdf')  # * means all files, if need specific extension then for example *.pdf
    
    # replace slash in the expiration date as it is a forbidden character when renaming the file
    replaced_expiry_date = stripped_expiry_date.replace("/", "")
    
    # RENAME DOWNLOADED FILE
    # locate downloaded certificate
    name_element = driver.find_element(By.XPATH, 
                    "/html/body/intact-root/div/intact-supplier-detail/div/mat-card/div/div[1]/div/intact-company-info/div/h2").text
    time.sleep(5)
    list_of_files = glob.glob(path)
    latest_file = max(list_of_files, key=os.path.getctime)    
    old_file_path = os.path.abspath(latest_file)
    # remove forbidden characters
    char_remove = ["/", "|", ":", "*", "?", "!", "<", ">"]
    for char in char_remove:
        if char in name_element:
            new_name = name_element.replace(char, "_")
            break
        else:
            new_name = name_element
    # create new name for the certificate
    new_name_path = str(user_path + '\\' + str(OIN) + '_' + new_name + '_' + replaced_expiry_date + '.pdf')
    replaced_new_name_path = str(new_name_path.replace("\n", ""))
    time.sleep(3)
    
    # MOVE DOWNLOADED FILE
    try:
        os.rename(old_file_path, replaced_new_name_path)
        shutil.move(replaced_new_name_path, destination_dir_path)
    except FileExistsError:
        # REMOVE OLD FILE, RENAME AND MOVE NEW FILE
        os.remove(replaced_new_name_path)
        os.rename(old_file_path, replaced_new_name_path)
        shutil.move(replaced_new_name_path, destination_dir_path)
    except:
        pass



def create_report(destination_dir_path):
    
    
    # SAVE SCRAPED DATA TO EXCEL
    
    df1 = pd.DataFrame(
        {
            'OIN' : OIN_found_list,
            'Name': None,
            'Department': None,
            'Suite': None,
            'Street': None,
            'City': None,
            'State': None,
            'Postal Code': None,
            'Country': None,
            'Cert Number': None,
            'Standard Series': None,
            'Standard': standard_results2,
            'Cert Issue Date': None,
            'Cert Expire Date': expiry_results2,
            'Issuing CB for the Latest Cert': None,
            'Certificate Status': status_results2,
            'Last Audit Visit Date': None,            
            'Certificate Scope': scope_results2
        }
    )
    
    df2 = pd.DataFrame(
        {
            'OIN' : OIN_found_list,
            'Name': None,
            'Type': None,
            'Department': None,
            'Suite': None,
            'Street': None,
            'City': None,
            'State': None,
            'Postal Code': None,
            'Country': None,
            'Supplier Gave Access To ARS': None,
            'Supplier Admin Email(s)': None,
            'Organization Email': email_results2
        }
    )
    
    df3 = pd.DataFrame(
        {
            'OIN' : OIN_not_found_list,
            'Standard': None,
            'Cert Expire Date': None,
            'Certificate Status': None
        }
    )

    # CREATE XLSX REPORT --> one file, 3 separate sheets

    date_string = datetime.now().strftime('%Y%m%d %H;%M;%S')
    with pd.ExcelWriter("REPORT_" + date_string +".xlsx") as writer:
        df1.to_excel(writer, sheet_name="OASIS", index=False)
        df2.to_excel(writer, sheet_name="OASIS Emails", index=False)
        df3.to_excel(writer, sheet_name="ALL OTHER STATUS T. CERTNOT FOD", index=False)

    # MOVE REPORT

    filename1 = "REPORT_" + date_string +".xlsx"
    filename_path1 = os.path.abspath(filename1)
    shutil.move(filename_path1, destination_dir_path)

    # CREATE ANOTHER DATAFRAME

    df4 = pd.DataFrame(
        {
            'OIN' : OIN_list,
            'Supplier name': supplier_name_results,
            'Supplier address': supplier_address_results,
            'Contact person': contact_results,
            'E-mail': email_results,
            'Status': status_results,
            'Series': series_results,
            'Standard': standard_results,
            'Latest certificate': latest_certificate_results,
            'Latest audit': latest_audit_results,
            'Expiry date': expiry_results,
            'Certificate Scope': scope_results,
        }
    )

    # CREATE XLSX REPORT --> one file, one sheet

    date_string = datetime.now().strftime('%Y%m%d %H;%M;%S')
    with pd.ExcelWriter("Compared Results_" + date_string +".xlsx") as writer:
        df4.to_excel(writer, sheet_name="Sheet1", index=False)


    # MOVE REPORT

    filename2 = "Compared Results_" + date_string +".xlsx"
    filename_path2 = os.path.abspath(filename2)
    shutil.move(filename_path2, destination_dir_path)
    
    print('ALL DONE! REPORT CREATED.')
    
    # INITIATE FUNCTIONS
    # locate_reports(destination_dir_path, filename2)
    # compare_reports(destination_dir_path, latest_report, previous_report)


    
def locate_reports(destination_dir_path, filename2):
    
    """
    Function needed to carry out the compare_reports() function.
    Locates reports which need to be compared.
    """
    
    global latest_report, previous_report
    
    # PATH TO ALL REPORTS
    path = r'\\components\DATA\PLS-Public\PLSQA\OASIS\OASIS_QMS certificates'
    #PATH TO LATEST REPORT
    latest_report = (destination_dir_path + '\\' + filename2)
    # GET SECOND LAST FOLDER AND REPORT PATH 
    folders_list = os.listdir(path)
    previous_folder = sorted(folders_list)[-4]
    previous_folder_path = (path + '\\' + previous_folder)
    for file in os.listdir(previous_folder_path):
        if file.startswith('Compared Results_'):
            previous_report = (previous_folder_path + '\\' + file)


    
def compare_reports(destination_dir_path, latest_report, previous_report):
    
    """
    Function used to compare just created report in the create_report() function 
    and the second last report to catch any differences such as
    supplier name/address change, whether the supplier lost/gained certification status, etc.
    """
    
    print(previous_report)
    print(latest_report)
    
    # READ EXCEL FILES
    df1 = pd.read_excel(previous_report, sheet_name='Sheet1', na_values=['NA'])
    df2 = pd.read_excel(latest_report, sheet_name='Sheet1', na_values=['NA'])
    
    # CREATE ANOTHER DATAFRAME
    df3 = pd.concat([df1,df2],sort=False)
    
    # REMOVE DUPLICATES, HIGHLUGHT CHANGES WITH <---
    df3a = df3.astype(str).groupby('OIN').agg(lambda x: '<---'.join(set(x)))
    
    # HIGHLUGHT CHANGES WITH COLOR
    df3b = df3a.style.map(lambda x: "background-color: red" if '<---' in x else "background-color: white")
 
    # SAVE FILE
    df3b.to_excel(latest_report)
       

    
    
    
    
    
    
